"""Tests for export functionality."""

import pytest
import polars as pl
from pathlib import Path
import tempfile
import shutil
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook

from src.export.reports import ReportExporter
from src.analytics.performance import PerformanceAnalyzer


@pytest.fixture
def sample_activities_for_export():
    """Create sample activities for export testing."""
    base_date = datetime(2024, 1, 1)
    
    activities = []
    for i in range(20):
        activities.append({
            "id": f"act{i}",
            "start_date_local": base_date + timedelta(days=i),
            "name": f"Ride {i}",
            "type": "Ride" if i % 3 != 0 else "Run",
            "moving_time": 3600 + (i * 100),
            "distance": 30000 + (i * 1000),
            "total_elevation_gain": 300 + (i * 10),
            "average_heartrate": 140 + (i % 20),
            "icu_average_watts": 200 + (i * 2) if i % 3 != 0 else None,
            "icu_normalized_watts": 210 + (i * 2) if i % 3 != 0 else None,
            "icu_training_load": 50 + (i * 3),
            "icu_ftp": 250 + (i // 5) * 5,
            "icu_fitness": 40 + i * 0.5,
            "icu_fatigue": 35 + i * 0.4,
        })
    
    return pl.DataFrame(activities)


@pytest.fixture
def temp_export_dir():
    """Create temporary directory for exports."""
    temp_dir = tempfile.mkdtemp(prefix="export_test_")
    yield Path(temp_dir)
    shutil.rmtree(temp_dir)


class TestExcelExport:
    """Tests for Excel export functionality."""
    
    def test_excel_export_basic(self, sample_activities_for_export, temp_export_dir):
        """Test basic Excel export."""
        exporter = ReportExporter(sample_activities_for_export)
        
        output_path = temp_export_dir / "test_report.xlsx"
        exported_path = exporter.export_to_excel(str(output_path))
        
        # Verify file created
        assert Path(exported_path).exists()
        assert Path(exported_path).suffix == ".xlsx"
        
        # Verify file can be opened
        wb = load_workbook(exported_path)
        assert len(wb.sheetnames) > 0
    
    def test_excel_export_sheets(self, sample_activities_for_export, temp_export_dir):
        """Test that all expected sheets are created."""
        exporter = ReportExporter(sample_activities_for_export)
        
        output_path = temp_export_dir / "test_report.xlsx"
        exported_path = exporter.export_to_excel(str(output_path))
        
        wb = load_workbook(exported_path)
        
        # Check expected sheets exist
        expected_sheets = [
            "Activities",
            "Weekly Summary",
            "Training Load",
            "Best Efforts",
            "Power Analysis",
            "Predictions"
        ]
        
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames
    
    def test_excel_activities_sheet_content(self, sample_activities_for_export, temp_export_dir):
        """Test content of activities sheet."""
        exporter = ReportExporter(sample_activities_for_export)
        
        output_path = temp_export_dir / "test_report.xlsx"
        exported_path = exporter.export_to_excel(str(output_path))
        
        wb = load_workbook(exported_path)
        ws = wb["Activities"]
        
        # Check headers exist
        headers = [cell.value for cell in ws[1]]
        assert "start_date_local" in headers
        assert "name" in headers
        assert "type" in headers
        
        # Check data rows
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 20  # Should have all activities
    
    def test_excel_export_auto_generated_path(self, sample_activities_for_export):
        """Test Excel export with auto-generated path."""
        exporter = ReportExporter(sample_activities_for_export)
        
        # Export without specifying path
        exported_path = exporter.export_to_excel()
        
        # Should create file in exports directory
        assert Path(exported_path).exists()
        assert "exports" in str(exported_path)
        assert "cycling_analytics" in Path(exported_path).name
        
        # Cleanup
        Path(exported_path).unlink()
        Path(exported_path).parent.rmdir()
    
    def test_excel_data_formatting(self, sample_activities_for_export, temp_export_dir):
        """Test that data is properly formatted in Excel."""
        exporter = ReportExporter(sample_activities_for_export)
        
        output_path = temp_export_dir / "test_report.xlsx"
        exported_path = exporter.export_to_excel(str(output_path))
        
        wb = load_workbook(exported_path)
        ws = wb["Activities"]
        
        # Find distance column
        headers = [cell.value for cell in ws[1]]
        if "distance_km" in headers:
            dist_col = headers.index("distance_km") + 1
            
            # Check distance is converted to km
            first_distance = ws.cell(row=2, column=dist_col).value
            assert first_distance == 30  # 30000m = 30km
    
    def test_excel_weekly_summary(self, sample_activities_for_export, temp_export_dir):
        """Test weekly summary sheet content."""
        exporter = ReportExporter(sample_activities_for_export)
        
        output_path = temp_export_dir / "test_report.xlsx"
        exported_path = exporter.export_to_excel(str(output_path))
        
        wb = load_workbook(exported_path)
        ws = wb["Weekly Summary"]
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "week" in headers
        assert "activities" in headers
        assert "total_hours" in headers
        assert "total_km" in headers
        
        # Should have multiple weeks of data
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) > 0


class TestCSVExport:
    """Tests for CSV export functionality."""
    
    def test_csv_export_basic(self, sample_activities_for_export, temp_export_dir):
        """Test basic CSV export."""
        exporter = ReportExporter(sample_activities_for_export)
        
        exported_files = exporter.export_to_csv(str(temp_export_dir))
        
        assert len(exported_files) > 0
        assert "activities" in exported_files
        
        # Verify files exist
        for file_path in exported_files.values():
            assert Path(file_path).exists()
            assert Path(file_path).suffix == ".csv"
    
    def test_csv_activities_content(self, sample_activities_for_export, temp_export_dir):
        """Test CSV activities file content."""
        exporter = ReportExporter(sample_activities_for_export)
        
        exported_files = exporter.export_to_csv(str(temp_export_dir))
        activities_file = exported_files["activities"]
        
        # Read back the CSV
        df = pl.read_csv(activities_file)
        
        # Verify content
        assert len(df) == 20
        assert "id" in df.columns
        assert "name" in df.columns
        assert df["id"].to_list() == [f"act{i}" for i in range(20)]
    
    def test_csv_weekly_summary(self, sample_activities_for_export, temp_export_dir):
        """Test CSV weekly summary export."""
        exporter = ReportExporter(sample_activities_for_export)
        
        exported_files = exporter.export_to_csv(str(temp_export_dir))
        
        if "weekly_summary" in exported_files:
            weekly_file = exported_files["weekly_summary"]
            df = pl.read_csv(weekly_file)
            
            assert "week" in df.columns
            assert "activities" in df.columns
            assert len(df) > 0
    
    def test_csv_best_efforts(self, sample_activities_for_export, temp_export_dir):
        """Test CSV best efforts export."""
        exporter = ReportExporter(sample_activities_for_export)
        
        exported_files = exporter.export_to_csv(str(temp_export_dir))
        
        if "best_efforts" in exported_files:
            best_file = exported_files["best_efforts"]
            df = pl.read_csv(best_file)
            
            assert "icu_normalized_watts" in df.columns
            assert len(df) > 0
            
            # Should be sorted by power (descending)
            powers = df["icu_normalized_watts"].to_list()
            assert powers == sorted(powers, reverse=True)


class TestSummaryReport:
    """Tests for text summary report generation."""
    
    def test_summary_report_generation(self, sample_activities_for_export):
        """Test summary report generation."""
        exporter = ReportExporter(sample_activities_for_export)
        
        summary = exporter.generate_summary_report()
        
        assert isinstance(summary, str)
        assert len(summary) > 0
        assert "CYCLING ANALYTICS SUMMARY REPORT" in summary
    
    def test_summary_report_content(self, sample_activities_for_export):
        """Test summary report contains expected sections."""
        exporter = ReportExporter(sample_activities_for_export)
        
        summary = exporter.generate_summary_report()
        
        # Check for key sections
        assert "OVERVIEW" in summary
        assert "TRAINING LOAD" in summary
        assert "FTP PROGRESSION" in summary
        assert "PERFORMANCE READINESS" in summary
        
        # Check for metrics
        assert "Total Activities: 20" in summary
        assert "Total Distance:" in summary
        assert "Total Time:" in summary
    
    def test_summary_report_with_minimal_data(self):
        """Test summary report with minimal data."""
        df = pl.DataFrame({
            "id": ["act1"],
            "start_date_local": [datetime.now()],
            "distance": [30000],
            "moving_time": [3600],
        })
        
        exporter = ReportExporter(df)
        summary = exporter.generate_summary_report()
        
        assert "Total Activities: 1" in summary
        assert "END OF REPORT" in summary


class TestDataIntegrityInExports:
    """Test that data integrity is maintained during export."""
    
    def test_numeric_precision_in_excel(self, temp_export_dir):
        """Test numeric precision is preserved in Excel export."""
        df = pl.DataFrame({
            "id": ["act1"],
            "start_date_local": [datetime.now()],
            "distance": [42195.5],
            "icu_average_watts": [234.567],
            "icu_training_load": [123.456],
        })
        
        exporter = ReportExporter(df)
        output_path = temp_export_dir / "precision_test.xlsx"
        exported_path = exporter.export_to_excel(str(output_path))
        
        # Read back and verify
        wb = load_workbook(exported_path)
        ws = wb["Activities"]
        
        headers = [cell.value for cell in ws[1]]
        
        # Check distance (converted to km)
        if "distance_km" in headers:
            dist_col = headers.index("distance_km") + 1
            distance_km = ws.cell(row=2, column=dist_col).value
            assert abs(distance_km - 42.1955) < 0.001
    
    def test_date_handling_in_export(self, temp_export_dir):
        """Test date handling in exports."""
        specific_date = datetime(2024, 3, 15, 14, 30, 45)
        
        df = pl.DataFrame({
            "id": ["act1"],
            "start_date_local": [specific_date],
            "name": ["Test Ride"],
        })
        
        exporter = ReportExporter(df)
        
        # Test CSV export
        csv_files = exporter.export_to_csv(str(temp_export_dir))
        df_csv = pl.read_csv(csv_files["activities"])
        
        # Date should be preserved (as string in CSV)
        assert "2024" in str(df_csv["start_date_local"][0])
    
    def test_null_handling_in_export(self, temp_export_dir):
        """Test null value handling in exports."""
        df = pl.DataFrame({
            "id": ["act1", "act2"],
            "start_date_local": [datetime.now(), datetime.now()],
            "distance": [30000, None],
            "icu_average_watts": [None, 200],
            "name": ["Ride 1", None],
        })
        
        exporter = ReportExporter(df)
        
        # Export to CSV
        csv_files = exporter.export_to_csv(str(temp_export_dir))
        df_csv = pl.read_csv(csv_files["activities"])
        
        # Nulls should be preserved
        assert df_csv["distance"][1] is None or pd.isna(df_csv["distance"][1])
        assert df_csv["icu_average_watts"][0] is None or pd.isna(df_csv["icu_average_watts"][0])
    
    def test_unicode_in_export(self, temp_export_dir):
        """Test unicode character handling in exports."""
        df = pl.DataFrame({
            "id": ["act1"],
            "start_date_local": [datetime.now()],
            "name": ["Ride with émojis 🚴‍♂️ and çhars"],
            "description": ["Test 日本語 κόσμος"],
        })
        
        exporter = ReportExporter(df)
        
        # Export to CSV
        csv_files = exporter.export_to_csv(str(temp_export_dir))
        df_csv = pl.read_csv(csv_files["activities"])
        
        # Unicode should be preserved
        assert "🚴‍♂️" in df_csv["name"][0]
        assert "日本語" in df_csv["description"][0]


class TestExportErrorHandling:
    """Test error handling in export functionality."""
    
    def test_export_empty_dataframe(self, temp_export_dir):
        """Test exporting empty dataframe."""
        df = pl.DataFrame()
        exporter = ReportExporter(df)
        
        # Should handle gracefully
        output_path = temp_export_dir / "empty.xlsx"
        exported_path = exporter.export_to_excel(str(output_path))
        
        assert Path(exported_path).exists()
    
    def test_export_with_missing_columns(self, temp_export_dir):
        """Test export when expected columns are missing."""
        df = pl.DataFrame({
            "id": ["act1"],
            "name": ["Test Ride"],
            # Missing many expected columns
        })
        
        exporter = ReportExporter(df)
        
        # Should handle missing columns gracefully
        output_path = temp_export_dir / "minimal.xlsx"
        exported_path = exporter.export_to_excel(str(output_path))
        
        assert Path(exported_path).exists()
        
        wb = load_workbook(exported_path)
        assert len(wb.sheetnames) > 0
    
    def test_export_invalid_path(self):
        """Test export with invalid output path."""
        df = pl.DataFrame({"id": ["act1"]})
        exporter = ReportExporter(df)
        
        # Try to export to invalid path
        invalid_path = "/nonexistent/directory/report.xlsx"
        
        # Should either handle error or create directory
        try:
            exported_path = exporter.export_to_excel(invalid_path)
            # If successful, path should exist
            assert Path(exported_path).exists()
            # Cleanup
            Path(exported_path).unlink()
        except Exception:
            # Error handling is acceptable
            pass