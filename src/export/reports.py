"""Export functionality for cycling analytics data and reports."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import polars as pl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from src.analytics.performance import PerformanceAnalyzer
from src.analytics.advanced import AdvancedAnalytics
from src.ml.predictions import PerformancePredictor

logger = logging.getLogger(__name__)


class ReportExporter:
    """Export cycling analytics data and reports."""
    
    def __init__(self, activities_df: pl.DataFrame):
        """Initialize report exporter.
        
        Args:
            activities_df: DataFrame with activity data
        """
        self.df = activities_df
        self.analyzer = PerformanceAnalyzer(activities_df)
        self.advanced = AdvancedAnalytics(activities_df)
        self.predictor = PerformancePredictor(activities_df)
    
    def export_to_excel(self, output_path: str = None) -> str:
        """Export comprehensive analytics to Excel.
        
        Args:
            output_path: Output file path (generated if not provided)
            
        Returns:
            Path to exported file
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"exports/cycling_analytics_{timestamp}.xlsx"
        
        # Ensure export directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Activities Summary Sheet
        self._add_activities_sheet(wb)
        
        # 2. Weekly Summary Sheet
        self._add_weekly_summary_sheet(wb)
        
        # 3. Training Load Sheet
        self._add_training_load_sheet(wb)
        
        # 4. Best Efforts Sheet
        self._add_best_efforts_sheet(wb)
        
        # 5. Power Analysis Sheet
        self._add_power_analysis_sheet(wb)
        
        # 6. Predictions Sheet
        self._add_predictions_sheet(wb)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report exported to {output_path}")
        
        return output_path
    
    def _add_activities_sheet(self, wb: Workbook):
        """Add activities summary sheet to workbook."""
        ws = wb.create_sheet("Activities")
        
        # Select relevant columns
        columns_to_export = [
            "start_date_local", "name", "type", "moving_time", "distance",
            "total_elevation_gain", "average_speed", "average_heartrate",
            "icu_average_watts", "icu_normalized_watts", "icu_training_load",
            "icu_intensity", "icu_ftp"
        ]
        
        # Filter columns that exist
        available_columns = [col for col in columns_to_export if col in self.df.columns]
        
        # Convert to pandas for Excel export
        df_export = self.df.select(available_columns).to_pandas()
        
        # Format columns
        if "moving_time" in df_export.columns:
            df_export["moving_time"] = df_export["moving_time"] / 3600  # Convert to hours
            df_export.rename(columns={"moving_time": "duration_hours"}, inplace=True)
        
        if "distance" in df_export.columns:
            df_export["distance"] = df_export["distance"] / 1000  # Convert to km
            df_export.rename(columns={"distance": "distance_km"}, inplace=True)
        
        # Write to sheet
        for r in dataframe_to_rows(df_export, index=False, header=True):
            ws.append(r)
        
        # Format header
        self._format_header(ws)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _add_weekly_summary_sheet(self, wb: Workbook):
        """Add weekly summary sheet to workbook."""
        ws = wb.create_sheet("Weekly Summary")
        
        # Get weekly summary
        weekly_df = self.analyzer.calculate_weekly_summary(weeks=52)
        
        if not weekly_df.is_empty():
            # Convert to pandas
            df_export = weekly_df.to_pandas()
            
            # Format columns
            df_export["week"] = pd.to_datetime(df_export["week"]).dt.strftime("%Y-%m-%d")
            
            # Write to sheet
            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws.append(r)
            
            # Format header
            self._format_header(ws)
            
            # Auto-adjust columns
            self._auto_adjust_columns(ws)
    
    def _add_training_load_sheet(self, wb: Workbook):
        """Add training load sheet to workbook."""
        ws = wb.create_sheet("Training Load")
        
        # Calculate metrics for different periods
        periods = [7, 14, 28, 42, 90]
        metrics_data = []
        
        for days in periods:
            metrics = self.analyzer.calculate_training_load(days=days)
            metrics["period_days"] = days
            metrics_data.append(metrics)
        
        # Convert to DataFrame
        df_export = pd.DataFrame(metrics_data)
        
        # Reorder columns
        cols = ["period_days", "activities", "total_time_hours", "total_load", "ctl", "atl", "tsb"]
        available_cols = [col for col in cols if col in df_export.columns]
        df_export = df_export[available_cols]
        
        # Write to sheet
        for r in dataframe_to_rows(df_export, index=False, header=True):
            ws.append(r)
        
        # Format header
        self._format_header(ws)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _add_best_efforts_sheet(self, wb: Workbook):
        """Add best efforts sheet to workbook."""
        ws = wb.create_sheet("Best Efforts")
        
        # Get best efforts for different metrics
        metrics = {
            "Normalized Power (W)": "icu_normalized_watts",
            "Average Power (W)": "icu_average_watts",
            "Training Load": "icu_training_load",
            "Distance (m)": "distance",
            "Duration (s)": "moving_time",
        }
        
        row_num = 1
        
        for title, metric in metrics.items():
            if metric in self.df.columns:
                # Add section title
                ws.cell(row=row_num, column=1, value=f"Top 10 - {title}")
                ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
                row_num += 1
                
                # Get best efforts
                best = self.analyzer.get_best_efforts(metric=metric, top_n=10)
                
                if not best.is_empty():
                    df_export = best.to_pandas()
                    
                    # Format distance and time columns
                    if "distance" in df_export.columns:
                        df_export["distance"] = df_export["distance"] / 1000  # km
                    if "moving_time" in df_export.columns:
                        df_export["moving_time"] = df_export["moving_time"] / 60  # minutes
                    
                    # Write data
                    for r in dataframe_to_rows(df_export, index=False, header=True):
                        ws.append(r)
                        row_num += 1
                    
                    # Add spacing
                    row_num += 2
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _add_power_analysis_sheet(self, wb: Workbook):
        """Add power analysis sheet to workbook."""
        ws = wb.create_sheet("Power Analysis")
        
        # FTP Progression
        ftp_data = self.analyzer.calculate_ftp_progression()
        
        if not ftp_data.is_empty():
            ws.append(["FTP Progression"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            
            df_export = ftp_data.to_pandas()
            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws.append(r)
        
        # Add spacing
        ws.append([])
        ws.append([])
        
        # Power Zones
        zone_data = self.analyzer.calculate_power_zones()
        
        if zone_data:
            ws.append([f"Power Zones (FTP: {zone_data.get('ftp', 'N/A')}W)"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            
            # Zone boundaries
            ws.append(["Zone", "Lower Bound (W)", "Upper Bound (W)", "Time (hours)"])
            
            if zone_data.get("zones"):
                for zone_name, (lower, upper) in zone_data["zones"].items():
                    zone_num = zone_name.split()[0]  # Extract zone number
                    time_key = f"{zone_num.lower()}_secs"
                    time_hours = zone_data.get("zone_times_hours", {}).get(time_key, 0)
                    
                    upper_str = f"{upper:.0f}" if upper != float("inf") else "Max"
                    ws.append([zone_name, f"{lower:.0f}", upper_str, f"{time_hours:.1f}"])
        
        # Add spacing
        ws.append([])
        ws.append([])
        
        # Power Curve
        power_curve = self.advanced.calculate_power_curve()
        
        if not power_curve.is_empty():
            ws.append(["Power Curve (95th percentile)"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            
            df_export = power_curve.to_pandas()
            for r in dataframe_to_rows(df_export, index=False, header=True):
                ws.append(r)
        
        # Format header cells
        self._format_header(ws)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _add_predictions_sheet(self, wb: Workbook):
        """Add predictions sheet to workbook."""
        ws = wb.create_sheet("Predictions")
        
        # FTP Predictions
        ftp_pred = self.predictor.predict_ftp_progression(days_ahead=30)
        
        if "predictions" in ftp_pred:
            ws.append(["FTP Progression Forecast (30 days)"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            
            ws.append([f"Current FTP: {ftp_pred.get('current_ftp', 'N/A')}W"])
            ws.append([f"Expected FTP in 30 days: {ftp_pred['predictions']['ftp_values'][-1]:.0f}W"])
            ws.append([f"Expected Gain: {ftp_pred.get('expected_gain', 0):.0f}W ({ftp_pred.get('expected_gain_percentage', 0):.1f}%)"])
            ws.append([f"Best Model: {ftp_pred.get('best_model', 'N/A')}"])
            
            ws.append([])
        
        # Performance Readiness
        readiness = self.predictor.predict_performance_readiness()
        
        if "overall_readiness" in readiness:
            ws.append(["Performance Readiness"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            
            ws.append([f"Overall Readiness: {readiness['overall_readiness']:.0f}%"])
            
            if readiness.get("readiness_factors"):
                ws.append(["Factor", "Score"])
                for factor, score in readiness["readiness_factors"].items():
                    ws.append([factor.title(), f"{score:.0f}%"])
            
            if readiness.get("recommendations"):
                ws.append([])
                ws.append(["Recommendations:"])
                for rec in readiness["recommendations"]:
                    ws.append([f"• {rec}"])
            
            ws.append([])
        
        # Season Forecast
        season = self.predictor.forecast_season_trajectory(months_ahead=3)
        
        if "forecasts" in season:
            ws.append(["Season Trajectory (3 months)"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            
            for metric, forecast in season["forecasts"].items():
                ws.append([f"{metric}: {forecast['trend']} trend"])
            
            if season.get("recommendations"):
                ws.append([])
                ws.append(["Season Recommendations:"])
                for rec in season["recommendations"]:
                    ws.append([f"• {rec}"])
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _format_header(self, ws):
        """Format header row with styling."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Find and format header rows
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            # Check if this looks like a header row
            if row[0].value and isinstance(row[0].value, str):
                # Check if it's a data header (not section title)
                if any(cell.value in ["date", "week", "name", "type", "Zone", "Factor"] for cell in row):
                    for cell in row:
                        if cell.value:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center")
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_to_csv(self, output_dir: str = "exports/csv") -> Dict[str, str]:
        """Export data to CSV files.
        
        Args:
            output_dir: Output directory for CSV files
            
        Returns:
            Dictionary with exported file paths
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        exported_files = {}
        
        # Export activities
        activities_file = output_path / f"activities_{timestamp}.csv"
        self.df.write_csv(activities_file)
        exported_files["activities"] = str(activities_file)
        
        # Export weekly summary
        weekly_df = self.analyzer.calculate_weekly_summary(weeks=52)
        if not weekly_df.is_empty():
            weekly_file = output_path / f"weekly_summary_{timestamp}.csv"
            weekly_df.write_csv(weekly_file)
            exported_files["weekly_summary"] = str(weekly_file)
        
        # Export best efforts
        best_power = self.analyzer.get_best_efforts("icu_normalized_watts", top_n=50)
        if not best_power.is_empty():
            best_file = output_path / f"best_efforts_{timestamp}.csv"
            best_power.write_csv(best_file)
            exported_files["best_efforts"] = str(best_file)
        
        logger.info(f"CSV files exported to {output_dir}")
        
        return exported_files
    
    def generate_summary_report(self) -> str:
        """Generate a text summary report.
        
        Returns:
            Summary report as string
        """
        report = []
        report.append("=" * 60)
        report.append("CYCLING ANALYTICS SUMMARY REPORT")
        report.append("=" * 60)
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")
        
        # Overview
        report.append("OVERVIEW")
        report.append("-" * 40)
        report.append(f"Total Activities: {len(self.df)}")
        
        if not self.df.is_empty():
            if "distance" in self.df.columns:
                total_distance = self.df["distance"].sum() / 1000
                report.append(f"Total Distance: {total_distance:,.0f} km")
            
            if "moving_time" in self.df.columns:
                total_time = self.df["moving_time"].sum() / 3600
                report.append(f"Total Time: {total_time:,.0f} hours")
            
            if "total_elevation_gain" in self.df.columns:
                total_elevation = self.df["total_elevation_gain"].sum()
                report.append(f"Total Elevation: {total_elevation:,.0f} m")
            
            # Date range
            if "start_date_local" in self.df.columns:
                min_date = self.df["start_date_local"].min()
                max_date = self.df["start_date_local"].max()
                report.append(f"Date Range: {min_date} to {max_date}")
        
        report.append("")
        
        # Training Load
        report.append("TRAINING LOAD (Last 42 days)")
        report.append("-" * 40)
        
        training_load = self.analyzer.calculate_training_load(days=42)
        if training_load:
            if training_load.get("ctl"):
                report.append(f"CTL (Fitness): {training_load['ctl']:.1f}")
            if training_load.get("atl"):
                report.append(f"ATL (Fatigue): {training_load['atl']:.1f}")
            if training_load.get("tsb"):
                report.append(f"TSB (Form): {training_load['tsb']:.1f}")
            if training_load.get("total_load"):
                report.append(f"Total Load: {training_load['total_load']:.0f}")
        
        report.append("")
        
        # FTP Progression
        report.append("FTP PROGRESSION")
        report.append("-" * 40)
        
        ftp_data = self.analyzer.calculate_ftp_progression()
        if not ftp_data.is_empty():
            latest_ftp = float(ftp_data["ftp_value"].tail(1)[0])
            first_ftp = float(ftp_data["ftp_value"].head(1)[0])

            report.append(f"Current FTP: {latest_ftp}W")
            report.append(f"Initial FTP: {first_ftp}W")
            report.append(f"Total Gain: {latest_ftp - first_ftp}W")
            report.append(f"Data Points: {len(ftp_data)}")
        
        report.append("")
        
        # Performance Readiness
        report.append("PERFORMANCE READINESS")
        report.append("-" * 40)
        
        readiness = self.predictor.predict_performance_readiness()
        if "overall_readiness" in readiness:
            report.append(f"Overall Readiness: {readiness['overall_readiness']:.0f}%")
            
            if readiness.get("recommendations"):
                report.append("\nRecommendations:")
                for rec in readiness["recommendations"]:
                    report.append(f"  • {rec}")
        
        report.append("")
        report.append("=" * 60)
        report.append("END OF REPORT")
        report.append("=" * 60)
        
        return "\n".join(report)